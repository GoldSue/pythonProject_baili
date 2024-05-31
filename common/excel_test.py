import os
import openpyxl


class excelutil:


    def get_object_path(self):
        file = (os.path.dirname(os.path.dirname(__file__)))
        filepath = os.path.join(file,"data")
        return filepath

    def read_excel(self,filename="login_data.xlsx"):
        filedata = os.path.join(self.get_object_path(),filename)
        wb = openpyxl.load_workbook(filedata)
        # 可以添加更多处理逻辑，例如：
        sheet = wb['login']
        print(sheet.max_row,sheet.max_column)
        all_list = []
        for rows in range(2,sheet.max_row+1):
            temp_list = []
            for cols in range(1,sheet.max_column+1):
                temp_list.append(sheet.cell(rows,cols).value)
            all_list.append(temp_list)

        print(all_list)


if __name__ == '__main__':
    eu = excelutil()  # 创建类的实例
    eu.get_object_path()  # 调用实例方法
    eu.read_excel()  # 调用实例方法
