import os
import openpyxl


class ExcelUtil:

    def get_object_path(self):
        return os.path.abspath(os.path.dirname(__file__)).split("common")[0]

    def read_excel(self):
        wb = openpyxl.load_workbook(self.get_object_path()+"/data/login_data.xlsx")
        sheet = wb["login"]

        print(sheet.max_row,sheet.max_column)
        #循环
        all_list = []
        for rows in range(2,sheet.max_row+1):
            temp_list = []
            for cols in range(1,sheet.max_column+1):
                temp_list.append(sheet.cell(rows,cols).value)
            all_list.append(temp_list)
        return all_list


        # all_list = []
        # for row in range(2, sheet.max_row + 1):
        #     row_data = [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
        #     all_list.append(row_data)




if __name__ == '__main__':
    ExcelUtil().read_excel()

